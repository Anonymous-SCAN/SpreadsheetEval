"""
recalc.py — Recompute every cell VALUE in an .xlsx workbook.

This is the ground-truth "calculator" used both when constructing reference
solutions and when grading agent output. Per the SpreadsheetBench v2 / cc.md
spec, verification compares *computed values* (LibreOffice recalculation), NOT
formula strings.

Two backends, tried in order:

  1. LibreOffice headless (`soffice`) — the canonical recalc engine used by
     SpreadsheetBench. Used automatically if `soffice`/`libreoffice` is on PATH
     (this is the case inside the Harbor Docker environment, where the
     Dockerfile installs it).

  2. `formulas` pure-python engine — a drop-in fallback for hosts without
     LibreOffice (e.g. the CI box used to author this benchmark). It parses the
     workbook, builds the dependency graph and evaluates every formula.

Both backends return the same structure: a dict mapping
    (sheet_name, "A1") -> normalized python value
so downstream comparison is backend-agnostic.
"""

from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
import datetime
from typing import Dict, Tuple, Any

import openpyxl


CellKey = Tuple[str, str]  # (sheet_name, cell_coordinate e.g. "B7")


# --------------------------------------------------------------------------- #
# Value normalization                                                          #
# --------------------------------------------------------------------------- #
def normalize_value(v: Any) -> Any:
    """Normalize a single cell value for OJ-style comparison.

    - numbers      -> rounded to 2 decimals (float)
    - dates/times  -> ISO 'YYYY-MM-DD' string (date normalization)
    - empty / ""   -> None  (empty string and truly-empty cell are equivalent)
    - errors/text  -> stripped string
    """
    if v is None:
        return None
    # bool is a subtype of int; treat explicitly so True/1 don't collide oddly
    if isinstance(v, bool):
        return bool(v)
    if isinstance(v, (int, float)):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return str(v)
        # NaN / inf guard
        if f != f:  # NaN
            return None
        return round(f, 2)
    if isinstance(v, (datetime.datetime, datetime.date)):
        # date normalization -> YYYY-MM-DD
        if isinstance(v, datetime.datetime):
            # if it carries a real time component keep it, else just the date
            if v.hour or v.minute or v.second:
                return v.strftime("%Y-%m-%d %H:%M:%S")
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        # numeric string? try to coerce so text-number vs number differences
        # in the *value* domain don't create false diffs after recalc.
        try:
            f = float(s.replace(",", ""))
            return round(f, 2)
        except ValueError:
            return s
    return str(v)


# --------------------------------------------------------------------------- #
# Backend detection                                                            #
# --------------------------------------------------------------------------- #
def _find_soffice() -> str | None:
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    # common absolute locations
    for p in ("/usr/bin/soffice", "/usr/bin/libreoffice",
              "/opt/libreoffice/program/soffice"):
        if os.path.exists(p):
            return p
    return None


# --------------------------------------------------------------------------- #
# Backend 1: LibreOffice headless                                              #
# --------------------------------------------------------------------------- #
def _recalc_libreoffice(path: str, soffice: str) -> Dict[CellKey, Any]:
    """Convert-with-recalc via LibreOffice, then read cached values."""
    with tempfile.TemporaryDirectory() as td:
        # LibreOffice recalculates formulas on load when converting.
        # Use a macro-free conversion to xlsx into a temp dir.
        env = dict(os.environ)
        env["HOME"] = td  # isolate LO profile
        cmd = [
            soffice, "--headless", "--calc", "--convert-to", "xlsx",
            "--outdir", td, path,
        ]
        subprocess.run(cmd, env=env, check=True,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                       timeout=180)
        out = os.path.join(td, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
        if not os.path.exists(out):
            raise RuntimeError("LibreOffice conversion produced no output")
        wb = openpyxl.load_workbook(out, data_only=True)
        return _read_values(wb)


# --------------------------------------------------------------------------- #
# Backend 2: pure-python `formulas`                                            #
# --------------------------------------------------------------------------- #
def _recalc_formulas(path: str) -> Dict[CellKey, Any]:
    import formulas  # lazy import
    import logging
    logging.getLogger("formulas").setLevel(logging.CRITICAL)
    logging.getLogger("schedula").setLevel(logging.CRITICAL)

    # First read structure with openpyxl to know every populated cell and its
    # static (non-formula) values.
    wb = openpyxl.load_workbook(path, data_only=False)
    result: Dict[CellKey, Any] = {}
    # `formulas` upper-cases sheet names; map UPPER -> real title so keys stay
    # consistent with the openpyxl-read static cells.
    sheet_case = {ws.title.upper(): ws.title for ws in wb.worksheets}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    result[(ws.title, cell.coordinate)] = None  # fill below
                else:
                    result[(ws.title, cell.coordinate)] = normalize_value(cell.value)

    # Now evaluate formulas.
    try:
        xl = formulas.ExcelModel().loads(path).finish()
        sol = xl.calculate()
    except Exception:
        # If the engine chokes, at least return the static grid.
        return result

    base = os.path.basename(path).upper()
    for key, val in sol.items():
        # keys look like "'[FILE.XLSX]SHEETNAME'!A1"
        try:
            sheet, coord = _parse_formulas_key(key, base)
        except Exception:
            continue
        if sheet is None:
            continue
        # map formulas' upper-cased sheet name back to the real title
        sheet = sheet_case.get(sheet.upper(), sheet)
        pyval = _unwrap_formulas_value(val)
        result[(sheet, coord)] = normalize_value(pyval)
    return result


def _parse_formulas_key(key: str, base: str):
    # e.g.  "'[T.XLSX]DATA'!A3"
    if "!" not in key:
        return None, None
    ref, coord = key.rsplit("!", 1)
    ref = ref.strip().strip("'")
    if ref.startswith("[") and "]" in ref:
        fname, sheet = ref[1:].split("]", 1)
    else:
        sheet = ref
    # strip any absolute markers
    coord = coord.replace("$", "")
    # only single-cell coords (skip ranges)
    if ":" in coord:
        return None, None
    return sheet, coord


def _unwrap_formulas_value(val):
    """`formulas` returns numpy arrays / Ranges; unwrap to a scalar."""
    try:
        import numpy as np
    except Exception:
        np = None
    v = val
    # formulas Ranges expose .value
    if hasattr(v, "value"):
        v = v.value
    if np is not None and isinstance(v, np.ndarray):
        flat = v.flatten()
        if flat.size == 0:
            return None
        v = flat[0]
    if isinstance(v, (list, tuple)):
        while isinstance(v, (list, tuple)) and len(v) > 0:
            v = v[0]
        if isinstance(v, (list, tuple)):
            return None
    if np is not None and isinstance(v, np.generic):
        v = v.item()
    # formulas error sentinels
    s = str(v)
    if s in ("#N/A", "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"):
        return s
    return v


# --------------------------------------------------------------------------- #
# Shared helper                                                                #
# --------------------------------------------------------------------------- #
def _read_values(wb) -> Dict[CellKey, Any]:
    result: Dict[CellKey, Any] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                result[(ws.title, cell.coordinate)] = normalize_value(cell.value)
    return result


# --------------------------------------------------------------------------- #
# Public API                                                                   #
# --------------------------------------------------------------------------- #
def recalc(path: str, backend: str | None = None) -> Dict[CellKey, Any]:
    """Recompute all cell values in `path`.

    backend: None (auto) | "libreoffice" | "formulas"
    Returns {(sheet, coord): normalized_value}.
    """
    if backend is None:
        backend = "libreoffice" if _find_soffice() else "formulas"
    if backend == "libreoffice":
        so = _find_soffice()
        if not so:
            raise RuntimeError("LibreOffice requested but not found")
        return _recalc_libreoffice(path, so)
    if backend == "formulas":
        return _recalc_formulas(path)
    raise ValueError(f"unknown backend {backend!r}")


def active_backend() -> str:
    return "libreoffice" if _find_soffice() else "formulas"


if __name__ == "__main__":
    import sys
    p = sys.argv[1]
    print(f"[recalc] backend={active_backend()}")
    vals = recalc(p)
    for k in sorted(vals)[:50]:
        print(k, "=>", vals[k])
    print(f"... {len(vals)} cells total")
