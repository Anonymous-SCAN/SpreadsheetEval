"""
build_lib.py — shared helpers for programmatically constructing rich, high-
complexity spreadsheet workbooks.

These implement the "complex element injection" dimensions from cc.md §3.2:
merged cells, distractor sheets, hidden rows/cols, data validation, conditional
formatting, number formats, styling. Every scenario builder composes them so the
generated workbooks look and behave like real business models (SpreadsheetBench
v2 style), not toy sheets.

Nothing here is manual — callers drive everything with a seeded `random.Random`,
so the whole pipeline is scalable (cc.md §1.4 / §9).
"""

from __future__ import annotations
import random
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ styling --
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.0%'
INT_FMT = '#,##0'


def title(ws, cell_range, text):
    """Merged title bar spanning cell_range (e.g. 'A1:F1')."""
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    c = ws[top_left]
    c.value = text
    c.fill = TITLE_FILL
    c.font = WHITE_BOLD
    c.alignment = CENTER


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill = HEADER_FILL
        c.font = WHITE_BOLD
        c.border = BORDER
        c.alignment = CENTER


def stamp(ws, coord, value, *, fmt=None, bold=False, border=True):
    c = ws[coord]
    c.value = value
    if fmt:
        c.number_format = fmt
    if bold:
        c.font = BOLD
    if border:
        c.border = BORDER
    return c


def add_distractor_sheet(wb, rng, name=None):
    """A plausible-but-irrelevant sheet (old/abandoned data or personal notes).
    Pure noise to raise the needle-in-haystack difficulty (cc.md §3.2(5))."""
    name = name or rng.choice([
        "旧版测算_废弃", "Notes_DoNotUse", "Archive_2019", "Scratch",
        "OldForecast_v1", "临时草稿",
    ])
    ws = wb.create_sheet(name)
    title(ws, "A1:D1", "⚠ DEPRECATED — 请勿引用此表 / do not reference")
    header_row(ws, 2, ["项目 Item", "金额 Amount", "备注 Note", "日期"])
    for r in range(3, 3 + rng.randint(8, 18)):
        stamp(ws, f"A{r}", rng.choice(["杂项", "预留", "调整", "Misc", "TBD"]))
        stamp(ws, f"B{r}", round(rng.uniform(1000, 90000), 2), fmt=CURRENCY_FMT)
        stamp(ws, f"C{r}", rng.choice(["", "old", "check later", "废弃"]))
    return ws


def hide_helper_rows(ws, rows):
    for r in rows:
        ws.row_dimensions[r].hidden = True


def add_dropdown(ws, cell_range, options):
    dv = DataValidation(type="list",
                        formula1='"%s"' % ",".join(options),
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_range_validation(ws, cell_range, lo, hi):
    dv = DataValidation(type="decimal", operator="between",
                        formula1=str(lo), formula2=str(hi), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_traffic_light(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))


def negative_red(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(color="FF0000")))


def col(n):
    return get_column_letter(n)
