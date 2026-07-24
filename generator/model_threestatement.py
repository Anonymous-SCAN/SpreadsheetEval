"""
model_threestatement.py — a parametric, fully-programmatic integrated
financial model builder (3-statement + DCF), in the style of SpreadsheetBench
v2 Financial Modeling tasks.

One builder, driven by (seed, n_years, difficulty), produces a deep multi-sheet
workbook whose formula cells form a long cross-sheet dependency chain:

    Assumptions ─► Revenue ─► Income Statement ─┬─► Cash Flow ─► Balance Sheet
                              PP&E schedule  ────┤        ▲
                              Working-capital ───┘        │
                              DCF  ◄──────────────────────┘

Design guarantees:
  * The balance sheet BALANCES BY CONSTRUCTION (Assets == Liab+Equity every
    period) — a built-in self-consistency check the agent's answer must satisfy.
  * No circular references (interest on BEGINNING debt), so the pure-python
    `formulas` recalc backend evaluates it deterministically.
  * Only functions verified to work in both LibreOffice and the `formulas`
    fallback are used (SUM, IF, MAX, SUMPRODUCT, power, cross-sheet refs).

Each formula cell is registered as a "solution target" carrying:
    - the CORRECT formula                       (used for reference / Generation answer)
    - a list of ERROR VARIANTS (wrong formulas) (used for Debugging injection)

`emit_targets()` returns them so the task assembler can build the three views
(reference / generation-initial / debugging-initial).
"""

from __future__ import annotations
import random
from dataclasses import dataclass, field
from typing import List, Dict, Any

import openpyxl
from openpyxl.utils import get_column_letter as CL

import build_lib as B


@dataclass
class Target:
    sheet: str
    coord: str
    correct: str                     # correct formula string, incl leading '='
    kind: str                        # semantic tag (for instruction / difficulty)
    errors: List[Dict[str, str]] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Error-variant synthesis                                                      #
# --------------------------------------------------------------------------- #
def _mk_errors(correct: str, ctx: Dict[str, str]) -> List[Dict[str, str]]:
    """Produce plausible wrong formulas from a correct one + context refs.

    ctx may provide named sibling refs the mistake can latch onto:
      prev  : previous-period cell (off-by-one / wrong-period bug)
      alt   : a semantically-tempting-but-wrong cell
    Every variant is a *value*-changing mistake so the verifier catches it.
    """
    variants: List[Dict[str, str]] = []
    f = correct

    # 1) sign flip on the first binary operator (- <-> +)
    if "-" in f[1:]:
        flipped = f[0] + f[1:].replace("-", "", 1).replace("+", "-", 0)
        flipped = flipped.replace("", "+")
        if flipped != f:
            variants.append({"formula": flipped, "error_type": "符号取反 sign-flip",
                             "cascade_effect": "改变本项数值，向下游净额传播"})
    elif "+" in f[1:]:
        flipped = f[0] + f[1:].replace("+", "-", 1)
        variants.append({"formula": flipped, "error_type": "符号取反 sign-flip",
                         "cascade_effect": "改变本项数值，向下游净额传播"})

    # 2) wrong-period reference (off-by-one column) if a prev ref exists
    if ctx.get("prev") and ctx.get("cur"):
        v = f.replace(ctx["cur"], ctx["prev"])
        if v != f:
            variants.append({"formula": v, "error_type": "引用偏移 off-by-one-period",
                             "cascade_effect": "错用上期数据，期间衔接断裂"})

    # 3) alternate wrong cell substitution
    if ctx.get("alt") and ctx.get("swap_from"):
        v = f.replace(ctx["swap_from"], ctx["alt"])
        if v != f:
            variants.append({"formula": v, "error_type": "引用错误 wrong-cell",
                             "cascade_effect": "引用了错误的科目，勾稽关系被破坏"})

    # 4) hard-coded number instead of formula (a classic modelling anti-pattern)
    variants.append({"formula": "=0", "error_type": "公式丢失 hardcoded-zero",
                     "cascade_effect": "该科目被清零，三表不平"})

    # de-dup, keep <=3
    seen, out = set(), []
    for v in variants:
        if v["formula"] in seen or v["formula"] == correct:
            continue
        seen.add(v["formula"])
        out.append(v)
        if len(out) >= 3:
            break
    # ensure at least one
    if not out:
        out.append({"formula": "=0", "error_type": "公式丢失 hardcoded-zero",
                    "cascade_effect": "该科目被清零"})
    return out


# --------------------------------------------------------------------------- #
# The model                                                                    #
# --------------------------------------------------------------------------- #
INDUSTRIES = ["SaaS", "制造", "零售", "医疗", "能源", "地产"]


def build(seed: int, n_years: int = 4, difficulty: str = "high",
          n_distractors: int = 2):
    """Return (workbook, targets, meta). Everything is deterministic in `seed`."""
    rng = random.Random(seed)
    wb = openpyxl.Workbook()

    industry = rng.choice(INDUSTRIES)
    start_year = rng.choice([2024, 2025])
    years = [start_year + i for i in range(n_years)]     # forecast years (t=1..N)
    # ---- assumptions (all inputs) ----
    a = {
        "base_rev": round(rng.uniform(8000, 40000), 2),
        "growth": [round(rng.uniform(0.12, 0.35), 4) for _ in range(n_years)],
        "gm": round(rng.uniform(0.35, 0.72), 4),         # gross margin
        "opex_pct": round(rng.uniform(0.14, 0.28), 4),
        "tax": round(rng.uniform(0.15, 0.30), 4),
        "capex_pct": round(rng.uniform(0.05, 0.12), 4),
        "dep_years": rng.choice([5, 7, 10]),
        "dso": rng.randint(35, 75),
        "dio": rng.randint(30, 90),
        "dpo": rng.randint(30, 70),
        "int_rate": round(rng.uniform(0.04, 0.09), 4),
        "wacc": round(rng.uniform(0.09, 0.14), 4),
        "tg": round(rng.uniform(0.02, 0.035), 4),        # terminal growth
        "shares": rng.randint(800, 5000),
        # opening balances (t=0)
        "cash0": round(rng.uniform(2000, 8000), 2),
        "ppe_gross0": round(rng.uniform(6000, 20000), 2),
        "accdep0": round(rng.uniform(1000, 4000), 2),
        "debt0": round(rng.uniform(3000, 12000), 2),
        "repay": [round(rng.uniform(300, 1200), 2) for _ in range(n_years)],
    }

    # -------- HOUSE CONVENTIONS ("twists", cc.md §6 difficulty escalation) ----
    # Each twist is a *stated* non-textbook accounting convention that deviates
    # from the default a model would apply on autopilot. All are written into
    # the Assumptions sheet as explicit, human-readable notes (so the task stays
    # fair and fully verifiable), but each one changes a DRIVER formula whose
    # effect cascades across every downstream sheet. Under all-or-nothing OJ
    # over 150+ cells, mis-applying even one convention fails the whole task.
    #
    # Crucially, every twist keeps the three statements internally consistent
    # (same driver cell feeds IS / CF / BS), so the balance sheet still balances
    # by construction regardless of the values.
    all_twists = ["dep_begin", "salvage", "days360", "avg_interest",
                  "fixed_opex", "gm_ramp", "exit_multiple", "midyear_disc",
                  "capex_floor"]
    n_tw = {"low": 0, "medium": 2, "high": 5, "extreme": 8}.get(difficulty, 5)
    chosen = set(rng.sample(all_twists, min(n_tw, len(all_twists))))
    tw = {k: (k in chosen) for k in all_twists}
    # twist parameters (also become Assumptions cells)
    a["salvage_pct"] = round(rng.uniform(0.05, 0.15), 4) if tw["salvage"] else 0.0
    a["gm_ramp"] = round(rng.uniform(0.01, 0.03), 4) if tw["gm_ramp"] else 0.0
    a["fixed_opex"] = round(rng.uniform(500, 2500), 2) if tw["fixed_opex"] else 0.0
    a["exit_mult"] = round(rng.uniform(6, 12), 2) if tw["exit_multiple"] else 0.0
    a["days_basis"] = 360 if tw["days360"] else 365

    targets: List[Target] = []

    # ================= Sheet: Assumptions =================
    S_A = "Assumptions"
    aws = wb.active
    aws.title = S_A
    B.title(aws, "A1:F1", f"关键假设 Key Assumptions — {industry}")
    B.header_row(aws, 2, ["参数 Parameter", "值 Value", "单位", "说明"])
    rowmap: Dict[str, str] = {}   # param name -> B-column cell
    r = 3
    def put(name, label, val, fmt=None, unit=""):
        nonlocal r
        B.stamp(aws, f"A{r}", label)
        B.stamp(aws, f"B{r}", val, fmt=fmt)
        B.stamp(aws, f"C{r}", unit)
        rowmap[name] = f"{S_A}!B{r}"
        r += 1
    put("base_rev", "基年收入 Base revenue", a["base_rev"], B.CURRENCY_FMT, "千元")
    put("gm", "毛利率 Gross margin", a["gm"], B.PCT_FMT)
    put("opex_pct", "费用率 Opex %", a["opex_pct"], B.PCT_FMT)
    put("tax", "所得税率 Tax rate", a["tax"], B.PCT_FMT)
    put("capex_pct", "资本开支率 Capex %", a["capex_pct"], B.PCT_FMT)
    put("dep_years", "折旧年限 Dep. years", a["dep_years"], unit="年")
    put("dso", "应收周转 DSO", a["dso"], unit="天")
    put("dio", "存货周转 DIO", a["dio"], unit="天")
    put("dpo", "应付周转 DPO", a["dpo"], unit="天")
    put("int_rate", "利率 Interest rate", a["int_rate"], B.PCT_FMT)
    put("wacc", "WACC", a["wacc"], B.PCT_FMT)
    put("tg", "永续增长 Terminal growth", a["tg"], B.PCT_FMT)
    put("shares", "股本 Shares", a["shares"], B.INT_FMT, "千股")
    put("cash0", "期初现金 Opening cash", a["cash0"], B.CURRENCY_FMT)
    put("ppe_gross0", "期初固定资产原值", a["ppe_gross0"], B.CURRENCY_FMT)
    put("accdep0", "期初累计折旧", a["accdep0"], B.CURRENCY_FMT)
    put("debt0", "期初有息负债", a["debt0"], B.CURRENCY_FMT)
    # twist parameter cells (only those active) — each needs a real cell to ref
    if tw["salvage"]:
        put("salvage_pct", "残值率 Salvage %", a["salvage_pct"], B.PCT_FMT)
    if tw["gm_ramp"]:
        put("gm_ramp", "毛利率年提升 GM ramp/yr", a["gm_ramp"], B.PCT_FMT)
    if tw["fixed_opex"]:
        put("fixed_opex", "固定费用 Fixed opex", a["fixed_opex"], B.CURRENCY_FMT)
    if tw["exit_multiple"]:
        put("exit_mult", "退出倍数 Exit EV/EBIT", a["exit_mult"], unit="x")
    # growth + repay are per-year rows
    growth_cells, repay_cells = [], []
    for i, y in enumerate(years):
        B.stamp(aws, f"A{r}", f"收入增速 {y} Growth")
        B.stamp(aws, f"B{r}", a["growth"][i], fmt=B.PCT_FMT)
        growth_cells.append(f"{S_A}!B{r}"); r += 1
    for i, y in enumerate(years):
        B.stamp(aws, f"A{r}", f"还本 {y} Debt repayment")
        B.stamp(aws, f"B{r}", a["repay"][i], fmt=B.CURRENCY_FMT)
        repay_cells.append(f"{S_A}!B{r}"); r += 1
    B.add_range_validation(aws, "B3:B%d" % (r - 1), -1, 100000)
    aws.column_dimensions["A"].width = 26

    # ---- house-convention notes (explicit, so the task is fair) ----
    note_lines = []
    if tw["dep_begin"]:
        note_lines.append("折旧按【期初原值】计提（非期末）：Depreciation = 期初原值 / 折旧年限。")
    if tw["salvage"]:
        note_lines.append("折旧考虑残值：Depreciation = (期初原值 × (1−残值率)) / 折旧年限。")
    if tw["days360"]:
        note_lines.append("营运资本天数基数采用 360 天（非 365）：AR/存货/AP 均按 /360 计算。")
    if tw["avg_interest"]:
        note_lines.append("利息按【期初与期末负债的平均】计提：Interest = (期初+期末)/2 × 利率。")
    if tw["fixed_opex"]:
        note_lines.append("期间费用 = 收入 × 费用率 + 固定费用（含一笔固定管理费）。")
    if tw["gm_ramp"]:
        note_lines.append("毛利率逐年提升：第 t 年毛利率 = 基准毛利率 + GM ramp ×(t−1)。")
    if tw["exit_multiple"]:
        note_lines.append("终值采用【退出倍数法】：TV = 末期EBIT × 退出倍数（非永续增长法）。")
    if tw["midyear_disc"]:
        note_lines.append("DCF 采用【期中折现】：折现因子 = 1/(1+WACC)^(t−0.5)。")
    if tw["capex_floor"]:
        note_lines.append("资本开支设下限：Capex = MAX(收入 × 资本开支率, 上一年折旧)（维持性资本开支）。")
    if note_lines:
        r += 1
        B.stamp(aws, f"A{r}", "★ 会计口径约定 House conventions（务必遵循）", bold=True)
        r += 1
        for ln in note_lines:
            aws.merge_cells(f"A{r}:F{r}")
            B.stamp(aws, f"A{r}", "• " + ln, border=False)
            r += 1
    meta_twists = {k: v for k, v in tw.items() if v}

    # helper: year columns on schedule sheets. Col B = t0(open), C.. = years
    def yc(t):                         # t: 0..n_years  -> column letter
        return CL(2 + t)               # B is col2

    # a generic schedule-sheet setup: labels col A, year headers row 2, data row3+
    def new_sheet(name, banner):
        ws = wb.create_sheet(name)
        last = CL(2 + n_years)
        B.title(ws, f"A1:{last}1", banner)
        B.stamp(ws, "A2", "项目 \\ 年度", bold=True)
        B.stamp(ws, "B2", "期初 t0", bold=True)
        for t in range(1, n_years + 1):
            B.stamp(ws, f"{yc(t)}2", years[t - 1], bold=True)
        ws.column_dimensions["A"].width = 30
        return ws

    def line_label(ws, row, text):
        B.stamp(ws, f"A{row}", text, bold=True)

    def reg(ws_name, coord, correct, kind, ctx=None):
        t = Target(ws_name, coord, correct, kind,
                   _mk_errors(correct, ctx or {}))
        targets.append(t)
        return coord

    # ================= Sheet: Revenue =================
    S_R = "Revenue"
    rws = new_sheet(S_R, "收入预测 Revenue Forecast")
    ROW_REV = 3
    line_label(rws, ROW_REV, "营业收入 Revenue")
    B.stamp(rws, f"B{ROW_REV}", a["base_rev"], fmt=B.CURRENCY_FMT)  # t0 given
    for t in range(1, n_years + 1):
        cur = f"{yc(t)}{ROW_REV}"
        prev = f"{yc(t-1)}{ROW_REV}"
        g = growth_cells[t - 1]
        correct = f"={prev}*(1+{g})"
        reg(S_R, cur, correct, "revenue-growth",
            {"cur": cur, "prev": prev})

    def rev(t):
        return f"{S_R}!{yc(t)}{ROW_REV}"

    # ================= Sheet: PPE / Depreciation =================
    S_P = "PPE"
    pws = new_sheet(S_P, "固定资产与折旧 PP&E & Depreciation")
    R_GBEG, R_CAPX, R_GEND, R_DEP, R_ADEP, R_NET = 3, 4, 5, 6, 7, 8
    for row, lbl in [(R_GBEG, "期初原值 Gross begin"), (R_CAPX, "资本开支 Capex"),
                     (R_GEND, "期末原值 Gross end"), (R_DEP, "本期折旧 Depreciation"),
                     (R_ADEP, "累计折旧 Accum. dep"), (R_NET, "净值 Net PP&E")]:
        line_label(pws, row, lbl)
    # opening column
    B.stamp(pws, f"B{R_GEND}", a["ppe_gross0"], fmt=B.CURRENCY_FMT)
    B.stamp(pws, f"B{R_ADEP}", a["accdep0"], fmt=B.CURRENCY_FMT)
    reg(S_P, f"B{R_NET}", f"=B{R_GEND}-B{R_ADEP}", "ppe-net-open")
    for t in range(1, n_years + 1):
        c, p = yc(t), yc(t - 1)
        reg(S_P, f"{c}{R_GBEG}", f"={p}{R_GEND}", "ppe-gross-begin",
            {"cur": f"{c}{R_GEND}", "prev": f"{p}{R_GEND}"})
        # capex: optional maintenance floor = MAX(rev*capex%, prior-year dep)
        if tw["capex_floor"] and t > 1:
            capx_f = f"=MAX({rev(t)}*{rowmap['capex_pct']},{yc(t-1)}{R_DEP})"
        else:
            capx_f = f"={rev(t)}*{rowmap['capex_pct']}"
        reg(S_P, f"{c}{R_CAPX}", capx_f, "capex",
            {"swap_from": rowmap['capex_pct'], "alt": rowmap['opex_pct']})
        reg(S_P, f"{c}{R_GEND}", f"={c}{R_GBEG}+{c}{R_CAPX}", "ppe-gross-end")
        # depreciation base: end-of-period (default) or beginning (twist);
        # optionally net of salvage.
        base = f"{c}{R_GBEG}" if tw["dep_begin"] else f"{c}{R_GEND}"
        if tw["salvage"]:
            dep_f = f"={base}*(1-{rowmap['salvage_pct']})/{rowmap['dep_years']}"
        else:
            dep_f = f"={base}/{rowmap['dep_years']}"
        reg(S_P, f"{c}{R_DEP}", dep_f, "depreciation",
            {"swap_from": base, "alt": f"{c}{R_GEND}" if tw["dep_begin"] else f"{c}{R_GBEG}"})
        reg(S_P, f"{c}{R_ADEP}", f"={p}{R_ADEP}+{c}{R_DEP}", "accum-dep",
            {"cur": f"{c}{R_DEP}", "prev": f"{p}{R_DEP}"})
        reg(S_P, f"{c}{R_NET}", f"={c}{R_GEND}-{c}{R_ADEP}", "ppe-net")

    def dep(t):
        return f"{S_P}!{yc(t)}{R_DEP}"
    def capex(t):
        return f"{S_P}!{yc(t)}{R_CAPX}"
    def netppe(t):
        return f"{S_P}!{yc(t)}{R_NET}"

    # ================= Sheet: Income Statement =================
    S_I = "IncomeStatement"
    iws = new_sheet(S_I, "利润表 Income Statement")
    R_REV, R_COGS, R_GP, R_OPEX, R_DA, R_EBIT, R_INT, R_EBT, R_TAX, R_NI = range(3, 13)
    labels_i = ["营业收入 Revenue", "营业成本 COGS", "毛利 Gross profit",
                "期间费用 Opex", "折旧摊销 D&A", "息税前利润 EBIT",
                "利息费用 Interest", "税前利润 EBT", "所得税 Tax", "净利润 Net income"]
    for row, lbl in zip(range(R_REV, R_NI + 1), labels_i):
        line_label(iws, row, lbl)
    for t in range(1, n_years + 1):
        c = yc(t)
        reg(S_I, f"{c}{R_REV}", f"={rev(t)}", "is-revenue")
        # GM ramp: effective gross margin rises each year (twist)
        if tw["gm_ramp"]:
            gm_expr = f"({rowmap['gm']}+{rowmap['gm_ramp']}*{t-1})"
        else:
            gm_expr = f"{rowmap['gm']}"
        reg(S_I, f"{c}{R_COGS}", f"={c}{R_REV}*(1-{gm_expr})", "cogs",
            {"swap_from": rowmap['gm'], "alt": rowmap['opex_pct']})
        reg(S_I, f"{c}{R_GP}", f"={c}{R_REV}-{c}{R_COGS}", "gross-profit")
        # opex: variable + optional fixed component (twist)
        if tw["fixed_opex"]:
            opex_f = f"={c}{R_REV}*{rowmap['opex_pct']}+{rowmap['fixed_opex']}"
        else:
            opex_f = f"={c}{R_REV}*{rowmap['opex_pct']}"
        reg(S_I, f"{c}{R_OPEX}", opex_f, "opex",
            {"swap_from": rowmap['opex_pct'], "alt": rowmap['capex_pct']})
        reg(S_I, f"{c}{R_DA}", f"={dep(t)}", "is-da")
        reg(S_I, f"{c}{R_EBIT}", f"={c}{R_GP}-{c}{R_OPEX}-{c}{R_DA}", "ebit")
        # interest on beginning debt (from debt schedule below) — placeholder ref
        # debt schedule sheet defined next; forward-ref by cell address
        # (we know its geometry)
        # ================= Debt schedule cells resolved below
    # Debt schedule must exist before interest -> build it, then fill interest
    S_D = "DebtSchedule"
    dws = new_sheet(S_D, "债务计划 Debt Schedule")
    R_DBEG, R_REPAY, R_DEND = 3, 4, 5
    line_label(dws, R_DBEG, "期初负债 Debt begin")
    line_label(dws, R_REPAY, "还本 Repayment")
    line_label(dws, R_DEND, "期末负债 Debt end")
    B.stamp(dws, f"B{R_DEND}", a["debt0"], fmt=B.CURRENCY_FMT)  # opening debt
    for t in range(1, n_years + 1):
        c, p = yc(t), yc(t - 1)
        reg(S_D, f"{c}{R_DBEG}", f"={p}{R_DEND}", "debt-begin",
            {"cur": f"{c}{R_DEND}", "prev": f"{p}{R_DEND}"})
        reg(S_D, f"{c}{R_REPAY}", f"={repay_cells[t-1]}", "debt-repay")
        reg(S_D, f"{c}{R_DEND}", f"={c}{R_DBEG}-{c}{R_REPAY}", "debt-end")

    def debt_begin(t):
        return f"{S_D}!{yc(t)}{R_DBEG}"
    def debt_end(t):
        return f"{S_D}!{yc(t)}{R_DEND}"
    def repay(t):
        return f"{S_D}!{yc(t)}{R_REPAY}"

    # now fill interest/EBT/tax/NI on IS using debt schedule
    for t in range(1, n_years + 1):
        c = yc(t)
        # interest base: beginning debt (default) or average of begin+end (twist)
        if tw["avg_interest"]:
            int_f = (f"=({debt_begin(t)}+{debt_end(t)})/2*{rowmap['int_rate']}")
        else:
            int_f = f"={debt_begin(t)}*{rowmap['int_rate']}"
        reg(S_I, f"{c}{R_INT}", int_f, "interest",
            {"swap_from": debt_begin(t), "alt": debt_end(t)})
        reg(S_I, f"{c}{R_EBT}", f"={c}{R_EBIT}-{c}{R_INT}", "ebt")
        reg(S_I, f"{c}{R_TAX}", f"=MAX({c}{R_EBT},0)*{rowmap['tax']}", "tax",
            {"swap_from": rowmap['tax'], "alt": rowmap['gm']})
        reg(S_I, f"{c}{R_NI}", f"={c}{R_EBT}-{c}{R_TAX}", "net-income")

    def IS(row, t):
        return f"{S_I}!{yc(t)}{row}"

    # ================= Sheet: Working Capital =================
    S_W = "WorkingCapital"
    wws = new_sheet(S_W, "营运资本 Working Capital")
    R_AR, R_INV, R_AP, R_NWC, R_DNWC = 3, 4, 5, 6, 7
    for row, lbl in [(R_AR, "应收账款 AR"), (R_INV, "存货 Inventory"),
                     (R_AP, "应付账款 AP"), (R_NWC, "净营运资本 NWC"),
                     (R_DNWC, "NWC变动 ΔNWC")]:
        line_label(wws, row, lbl)
    # opening working capital (t0): derive from base rev/cogs to keep it realistic
    ar0 = round(a["base_rev"] * a["dso"] / 365, 2)
    cogs0 = a["base_rev"] * (1 - a["gm"])
    inv0 = round(cogs0 * a["dio"] / 365, 2)
    ap0 = round(cogs0 * a["dpo"] / 365, 2)
    B.stamp(wws, f"B{R_AR}", ar0, fmt=B.CURRENCY_FMT)
    B.stamp(wws, f"B{R_INV}", inv0, fmt=B.CURRENCY_FMT)
    B.stamp(wws, f"B{R_AP}", ap0, fmt=B.CURRENCY_FMT)
    reg(S_W, f"B{R_NWC}", f"=B{R_AR}+B{R_INV}-B{R_AP}", "nwc-open")
    for t in range(1, n_years + 1):
        c, p = yc(t), yc(t - 1)
        basis = a["days_basis"]   # 360 (twist) or 365 (default)
        reg(S_W, f"{c}{R_AR}", f"={IS(R_REV,t)}*{rowmap['dso']}/{basis}", "ar",
            {"swap_from": rowmap['dso'], "alt": rowmap['dio']})
        reg(S_W, f"{c}{R_INV}", f"={IS(R_COGS,t)}*{rowmap['dio']}/{basis}", "inventory",
            {"swap_from": rowmap['dio'], "alt": rowmap['dpo']})
        reg(S_W, f"{c}{R_AP}", f"={IS(R_COGS,t)}*{rowmap['dpo']}/{basis}", "ap",
            {"swap_from": rowmap['dpo'], "alt": rowmap['dso']})
        reg(S_W, f"{c}{R_NWC}", f"={c}{R_AR}+{c}{R_INV}-{c}{R_AP}", "nwc")
        reg(S_W, f"{c}{R_DNWC}", f"={c}{R_NWC}-{p}{R_NWC}", "delta-nwc",
            {"cur": f"{c}{R_NWC}", "prev": f"{p}{R_NWC}"})

    def WC(row, t):
        return f"{S_W}!{yc(t)}{row}"

    # ================= Sheet: Cash Flow (indirect) =================
    S_C = "CashFlow"
    cws = new_sheet(S_C, "现金流量表（间接法）Cash Flow (Indirect)")
    R_NI2, R_DA2, R_DNWC2, R_CFO, R_CAPX2, R_CFI, R_REPAY2, R_CFF, R_NETCF, R_CBEG, R_CEND = range(3, 14)
    labels_c = ["净利润 Net income", "加:折旧摊销 +D&A", "减:ΔNWC -ΔNWC",
                "经营活动现金流 CFO", "资本开支 Capex", "投资活动现金流 CFI",
                "偿还债务 Debt repay", "筹资活动现金流 CFF", "现金净变动 Net Δcash",
                "期初现金 Cash begin", "期末现金 Cash end"]
    for row, lbl in zip(range(R_NI2, R_CEND + 1), labels_c):
        line_label(cws, row, lbl)
    B.stamp(cws, f"B{R_CEND}", a["cash0"], fmt=B.CURRENCY_FMT)  # opening cash
    for t in range(1, n_years + 1):
        c, p = yc(t), yc(t - 1)
        reg(S_C, f"{c}{R_NI2}", f"={IS(R_NI,t)}", "cf-ni")
        reg(S_C, f"{c}{R_DA2}", f"={dep(t)}", "cf-da")
        reg(S_C, f"{c}{R_DNWC2}", f"=-{WC(R_DNWC,t)}", "cf-dnwc",
            {"swap_from": WC(R_DNWC, t), "alt": WC(R_NWC, t)})
        reg(S_C, f"{c}{R_CFO}", f"={c}{R_NI2}+{c}{R_DA2}+{c}{R_DNWC2}", "cfo")
        reg(S_C, f"{c}{R_CAPX2}", f"=-{capex(t)}", "cf-capex")
        reg(S_C, f"{c}{R_CFI}", f"={c}{R_CAPX2}", "cfi")
        reg(S_C, f"{c}{R_REPAY2}", f"=-{repay(t)}", "cf-repay")
        reg(S_C, f"{c}{R_CFF}", f"={c}{R_REPAY2}", "cff")
        reg(S_C, f"{c}{R_NETCF}", f"={c}{R_CFO}+{c}{R_CFI}+{c}{R_CFF}", "net-cf")
        reg(S_C, f"{c}{R_CBEG}", f"={p}{R_CEND}", "cash-begin",
            {"cur": f"{c}{R_CEND}", "prev": f"{p}{R_CEND}"})
        reg(S_C, f"{c}{R_CEND}", f"={c}{R_CBEG}+{c}{R_NETCF}", "cash-end")

    def CF(row, t):
        return f"{S_C}!{yc(t)}{row}"

    # ================= Sheet: Balance Sheet (+ balance check) =================
    S_B = "BalanceSheet"
    bws = new_sheet(S_B, "资产负债表 Balance Sheet")
    (R_CASH, R_ARB, R_INVB, R_NPPE, R_TA, R_APB, R_DEBTB, R_TL,
     R_EQ, R_TLE, R_CHK) = range(3, 14)
    labels_b = ["现金 Cash", "应收 AR", "存货 Inventory", "净固定资产 Net PP&E",
                "资产合计 Total assets", "应付 AP", "有息负债 Debt",
                "负债合计 Total liab", "所有者权益 Equity", "负债+权益 Total L&E",
                "平衡校验 Check(=0)"]
    for row, lbl in zip(range(R_CASH, R_CHK + 1), labels_b):
        line_label(bws, row, lbl)
    # opening equity = plug so opening BS balances
    npe0 = a["ppe_gross0"] - a["accdep0"]
    ta0 = a["cash0"] + ar0 + inv0 + npe0
    eq0 = round(ta0 - ap0 - a["debt0"], 2)
    B.stamp(bws, f"B{R_CASH}", a["cash0"], fmt=B.CURRENCY_FMT)
    B.stamp(bws, f"B{R_ARB}", ar0, fmt=B.CURRENCY_FMT)
    B.stamp(bws, f"B{R_INVB}", inv0, fmt=B.CURRENCY_FMT)
    B.stamp(bws, f"B{R_NPPE}", npe0, fmt=B.CURRENCY_FMT)
    B.stamp(bws, f"B{R_APB}", ap0, fmt=B.CURRENCY_FMT)
    B.stamp(bws, f"B{R_DEBTB}", a["debt0"], fmt=B.CURRENCY_FMT)
    B.stamp(bws, f"B{R_EQ}", eq0, fmt=B.CURRENCY_FMT)
    reg(S_B, f"B{R_TA}", f"=B{R_CASH}+B{R_ARB}+B{R_INVB}+B{R_NPPE}", "ta-open")
    reg(S_B, f"B{R_TL}", f"=B{R_APB}+B{R_DEBTB}", "tl-open")
    reg(S_B, f"B{R_TLE}", f"=B{R_TL}+B{R_EQ}", "tle-open")
    reg(S_B, f"B{R_CHK}", f"=B{R_TA}-B{R_TLE}", "check-open")
    for t in range(1, n_years + 1):
        c, p = yc(t), yc(t - 1)
        reg(S_B, f"{c}{R_CASH}", f"={CF(R_CEND,t)}", "bs-cash")
        reg(S_B, f"{c}{R_ARB}", f"={WC(R_AR,t)}", "bs-ar")
        reg(S_B, f"{c}{R_INVB}", f"={WC(R_INV,t)}", "bs-inv")
        reg(S_B, f"{c}{R_NPPE}", f"={netppe(t)}", "bs-nppe")
        reg(S_B, f"{c}{R_TA}", f"={c}{R_CASH}+{c}{R_ARB}+{c}{R_INVB}+{c}{R_NPPE}", "total-assets")
        reg(S_B, f"{c}{R_APB}", f"={WC(R_AP,t)}", "bs-ap")
        reg(S_B, f"{c}{R_DEBTB}", f"={debt_end(t)}", "bs-debt")
        reg(S_B, f"{c}{R_TL}", f"={c}{R_APB}+{c}{R_DEBTB}", "total-liab")
        reg(S_B, f"{c}{R_EQ}", f"={p}{R_EQ}+{IS(R_NI,t)}", "equity",
            {"cur": f"{p}{R_EQ}", "prev": f"{p}{R_EQ}"})
        reg(S_B, f"{c}{R_TLE}", f"={c}{R_TL}+{c}{R_EQ}", "total-le")
        reg(S_B, f"{c}{R_CHK}", f"={c}{R_TA}-{c}{R_TLE}", "balance-check")
    B.negative_red(bws, f"B{R_CHK}:{yc(n_years)}{R_CHK}")

    # ================= Sheet: DCF =================
    S_DCF = "DCF"
    fws = new_sheet(S_DCF, "现金流折现估值 DCF Valuation")
    R_EBITX, R_NOPAT, R_DAX, R_CAPXX, R_DNWCX, R_FCF, R_DF, R_PV = range(3, 11)
    labels_f = ["EBIT", "NOPAT=EBIT×(1-tax)", "加:D&A", "减:Capex", "减:ΔNWC",
                "无杠杆自由现金流 FCFF", "折现因子 DF", "FCF现值 PV"]
    for row, lbl in zip(range(R_EBITX, R_PV + 1), labels_f):
        line_label(fws, row, lbl)
    for t in range(1, n_years + 1):
        c = yc(t)
        reg(S_DCF, f"{c}{R_EBITX}", f"={IS(R_EBIT,t)}", "dcf-ebit")
        reg(S_DCF, f"{c}{R_NOPAT}", f"={c}{R_EBITX}*(1-{rowmap['tax']})", "nopat",
            {"swap_from": rowmap['tax'], "alt": rowmap['gm']})
        reg(S_DCF, f"{c}{R_DAX}", f"={dep(t)}", "dcf-da")
        reg(S_DCF, f"{c}{R_CAPXX}", f"=-{capex(t)}", "dcf-capex")
        reg(S_DCF, f"{c}{R_DNWCX}", f"=-{WC(R_DNWC,t)}", "dcf-dnwc")
        reg(S_DCF, f"{c}{R_FCF}", f"={c}{R_NOPAT}+{c}{R_DAX}+{c}{R_CAPXX}+{c}{R_DNWCX}", "fcff")
        # discount factor: end-of-period (default) or mid-year convention (twist)
        expo = f"({t}-0.5)" if tw["midyear_disc"] else f"{t}"
        reg(S_DCF, f"{c}{R_DF}", f"=1/(1+{rowmap['wacc']})^{expo}", "discount-factor")
        reg(S_DCF, f"{c}{R_PV}", f"={c}{R_FCF}*{c}{R_DF}", "pv-fcf")
    # valuation block (rows below)
    R_SUMPV, R_TV, R_PVTV, R_EV, R_NETDEBT, R_EQV, R_PS = range(R_PV + 2, R_PV + 9)
    lastcol = yc(n_years)
    fcf_range = f"{yc(1)}{R_FCF}:{lastcol}{R_FCF}"
    pv_range = f"{yc(1)}{R_PV}:{lastcol}{R_PV}"
    line_label(fws, R_SUMPV, "预测期PV合计 Σ PV(FCF)")
    reg(S_DCF, f"B{R_SUMPV}", f"=SUM({pv_range})", "sum-pv")
    line_label(fws, R_TV, "终值 Terminal value")
    # terminal value: Gordon growth (default) or exit-multiple on EBIT (twist)
    if tw["exit_multiple"]:
        tv_f = f"={lastcol}{R_EBITX}*{rowmap['exit_mult']}"
        tv_ctx = {"swap_from": rowmap['exit_mult'], "alt": rowmap['wacc']}
    else:
        tv_f = (f"={lastcol}{R_FCF}*(1+{rowmap['tg']})"
                f"/({rowmap['wacc']}-{rowmap['tg']})")
        tv_ctx = {"swap_from": rowmap['tg'], "alt": rowmap['wacc']}
    reg(S_DCF, f"B{R_TV}", tv_f, "terminal-value", tv_ctx)
    line_label(fws, R_PVTV, "终值现值 PV(TV)")
    reg(S_DCF, f"B{R_PVTV}", f"=B{R_TV}*{lastcol}{R_DF}", "pv-tv")
    line_label(fws, R_EV, "企业价值 Enterprise value")
    reg(S_DCF, f"B{R_EV}", f"=B{R_SUMPV}+B{R_PVTV}", "enterprise-value")
    line_label(fws, R_NETDEBT, "净负债 Net debt")
    reg(S_DCF, f"B{R_NETDEBT}", f"={rowmap['debt0']}-{rowmap['cash0']}", "net-debt")
    line_label(fws, R_EQV, "股权价值 Equity value")
    reg(S_DCF, f"B{R_EQV}", f"=B{R_EV}-B{R_NETDEBT}", "equity-value")
    line_label(fws, R_PS, "每股价值 Value per share")
    reg(S_DCF, f"B{R_PS}", f"=B{R_EQV}/{rowmap['shares']}", "value-per-share")

    # ================= distractors + hidden helper =================
    for i in range(n_distractors):
        B.add_distractor_sheet(wb, rng)
    # hide an internal helper row on PPE to force reasoning about hidden data
    if difficulty in ("high", "extreme"):
        pws.row_dimensions[R_ADEP].hidden = True

    meta = {
        "industry": industry, "years": years, "n_years": n_years,
        "difficulty": difficulty, "assumptions": a,
        "sheets": [ws.title for ws in wb.worksheets],
        "balance_check_cells":
            [f"{S_B}!{yc(t)}{R_CHK}" for t in range(0, n_years + 1)],
        "headline_cells": {
            "value_per_share": f"{S_DCF}!B{R_PS}",
            "enterprise_value": f"{S_DCF}!B{R_EV}",
        },
        "twists": sorted(meta_twists.keys()),
    }
    return wb, targets, meta


if __name__ == "__main__":
    wb, targets, meta = build(seed=7, n_years=4)
    print("sheets:", meta["sheets"])
    print("targets:", len(targets))
    wb.save("/tmp/model.xlsx")
    print("saved /tmp/model.xlsx")
