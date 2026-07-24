"""
model_agent.py — run a frontier model (Kimi-k3) as the solving agent on a
SpreadsheetEval task, to measure Overall Pass@1 for difficulty calibration
(cc.md §1.2 / §6).

Protocol (a fair, fully-automated agent harness):
  1. Serialize the agent's starting workbook (every populated cell, showing the
     formula or literal, plus a sheet/geometry map). This is the same
     information a human analyst opening the file would have.
  2. Send the task instruction + serialized workbook to the model, asking it to
     return a JSON list of cell edits: [{sheet, cell, formula}, ...].
  3. Apply the edits to a copy of the initial workbook -> output.xlsx.
  4. Recalc + grade with the exact OJ verifier used by Harbor.

Pass@1 over the suite = mean reward.
"""
from __future__ import annotations
import os
import io
import json
import argparse

import openpyxl

import recalc
import verifier_lib as V
import openrouter_client as OR

EVAL_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "SpreadsheetEval")

AGENT_SYS = (
    "You are an expert financial-modeling analyst working in Excel. You are given a "
    "multi-sheet workbook and a task. You must return the exact cell edits needed to "
    "complete the task. Use cross-sheet formula references (e.g. =IncomeStatement!C5) "
    "rather than hard-coded numbers wherever a value is derived. "
    "Return ONLY a JSON object of the form "
    '{\"edits\":[{\"sheet\":\"SheetName\",\"cell\":\"C5\",\"formula\":\"=A1+B1\"}, ...]}. '
    "Each edit sets one cell. Include every cell you need to create or fix. "
    "Do not include cells that should stay unchanged. No prose, only JSON."
)


def serialize_workbook(path, max_cells=4000):
    wb = openpyxl.load_workbook(path, data_only=False)
    lines = []
    for ws in wb.worksheets:
        hidden_note = ""
        rows = [r for r in ws.row_dimensions if ws.row_dimensions[r].hidden]
        if rows:
            hidden_note = f" (hidden rows: {sorted(rows)})"
        lines.append(f"### Sheet: {ws.title}{hidden_note}")
        n = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                v = c.value
                lines.append(f"{c.coordinate}: {v!r}")
                n += 1
                if n > max_cells:
                    break
    return "\n".join(lines)


def run_task(task_id, model=None, temperature=0.2):
    tdir = os.path.join(EVAL_DIR, task_id)
    init_path = os.path.join(tdir, "tests", "initial.xlsx")
    ref_path = os.path.join(tdir, "tests", "reference.xlsx")
    with open(os.path.join(tdir, "instruction.md")) as f:
        instruction = f.read()

    serial = serialize_workbook(init_path)
    user = (f"# Task instruction\n{instruction}\n\n"
            f"# Current workbook contents (cell: value/formula)\n{serial}\n\n"
            f"Return the JSON edits to produce output.xlsx.")

    reply = OR.chat([{"role": "system", "content": AGENT_SYS},
                     {"role": "user", "content": user}],
                    model=model, temperature=temperature,
                    max_tokens=28000, retries=1, timeout=600, no_fallback=True)
    try:
        obj = OR.extract_json(reply)
        edits = obj.get("edits", obj) if isinstance(obj, dict) else obj
    except Exception as e:
        return {"task_id": task_id, "reward": 0, "error": f"parse: {e}",
                "n_edits": 0}

    # apply edits to a fresh copy of initial
    wb = openpyxl.load_workbook(init_path, data_only=False)
    applied = 0
    for e in edits:
        try:
            sh = e["sheet"]; cell = e["cell"]; formula = e["formula"]
            if sh in wb.sheetnames:
                wb[sh][cell] = formula
                applied += 1
        except Exception:
            continue
    out = os.path.join(tdir, "tests", "_agent_output.xlsx")
    wb.save(out)

    init_vals = recalc.recalc(init_path)
    ref_vals = recalc.recalc(ref_path)
    out_vals = recalc.recalc(out)
    reward, detail = V.grade(init_vals, ref_vals, out_vals)
    return {"task_id": task_id, "reward": reward, "n_edits": applied,
            "modset": detail["modification_set_size"],
            "wrong": detail["wrong_modified_count"],
            "unexpected": detail["unexpected_change_count"],
            "sample_wrong": detail["wrong_modified"][:5]}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tasks", nargs="*", help="task ids (default: all)")
    ap.add_argument("--model", default=None)
    ap.add_argument("--out", default=os.path.join(
        os.path.dirname(EVAL_DIR), "model_eval.json"))
    args = ap.parse_args()

    if args.tasks:
        task_ids = args.tasks
    else:
        task_ids = sorted(d for d in os.listdir(EVAL_DIR)
                          if os.path.isdir(os.path.join(EVAL_DIR, d)))
    results = []
    for tid in task_ids:
        try:
            r = run_task(tid, model=args.model)
        except Exception as e:
            r = {"task_id": tid, "reward": 0, "error": str(e)[:200]}
        results.append(r)
        print(f"{tid:20s} reward={r['reward']} "
              f"edits={r.get('n_edits','-')} wrong={r.get('wrong','-')} "
              f"unexpected={r.get('unexpected','-')} {r.get('error','')}")
    passk = sum(r["reward"] for r in results) / max(len(results), 1)
    print(f"\nOverall Pass@1 = {passk:.3f}  ({sum(r['reward'] for r in results)}/{len(results)})")
    with open(args.out, "w") as f:
        json.dump({"model": args.model or OR.PRIMARY_MODEL,
                   "pass_at_1": passk, "results": results},
                  f, ensure_ascii=False, indent=2)
    print("wrote", args.out)


if __name__ == "__main__":
    main()
